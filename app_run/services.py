import io
import logging
from typing import Any

from django.db.models import QuerySet
from django.db.models.aggregates import Min, Max
from geopy.distance import geodesic
from openpyxl import load_workbook

from app_run.models import Position
from app_run.serializers import CollectibleItemSerializer

# Создаём логгер для этого модуля
logger = logging.getLogger(__name__)


def calculate_run_time_seconds(run) -> int:

    # Получить список позиций из модели Position для текущего забега с агрегацией по полю date_time
    # с минимальным и максимальным значением
    aggregates_date_time = Position.objects.filter(run=run).aggregate(
        min_date_time=Min("date_time"),
        max_date_time=Max("date_time"),
    )

    min_date_time = aggregates_date_time["min_date_time"]
    max_date_time = aggregates_date_time["max_date_time"]

    if min_date_time is None or max_date_time is None:
        return 0

    logger.warning(f"min_date_time: {min_date_time}")
    logger.warning(f"max_date_time: {max_date_time}")

    duration = (max_date_time - min_date_time).total_seconds()
    logger.warning(f"Затраченное время в секундах: {int(duration)}")

    return int(duration)


def calculate_route_distance(coordinates: QuerySet[Position, dict[str, Any]]) -> float:
    """
    Рассчитывает общее расстояние маршрута между географическими точками.

    Args:
        coordinates (list[dict]): Список координат в формате словарей,
                                  где каждый словарь содержит ключи "lat" и "lon".
                                  Пример: [{"lat": 55.7558, "lon": 37.6176}, ...]

    Returns:
        float: Общее расстояние маршрута в километрах.

    Raises:
        KeyError: Если в одном из элементов списка отсутствуют ключи "lat" или "lon".
        TypeError: Если значения "lat" или "lon" не являются числами.
        ValueError: Если список содержит менее двух точек.

    """
    if len(coordinates) < 2:
        raise ValueError(
            "Список должен содержать минимум две точки для расчета расстояния."
        )

    total_distance = 0.0

    for i in range(len(coordinates) - 1):
        point1 = (coordinates[i]["latitude"], coordinates[i]["longitude"])
        point2 = (coordinates[i + 1]["latitude"], coordinates[i + 1]["longitude"])
        distance = geodesic(point1, point2).kilometers
        total_distance += distance

    return total_distance


def read_excel_file(uploaded_file):
    # Читаем переданный файл
    file_content = uploaded_file.read()
    # Переводим файл в байтовый поток
    file_bytes = io.BytesIO(file_content)

    # Загружаем файл в экземпляр класса Workbook
    wb = load_workbook(file_bytes)
    # Получаем активный лист
    worksheet = wb.active
    # Получаем количество строк в листе
    max_row = worksheet.max_row

    # Создаем пустой список для хранения не валидных данных
    data_error = []

    # Получаем данные из листа
    for row in worksheet.iter_rows(values_only=True, min_row=2, max_row=max_row):
        # Присваиваем значения переменным
        name, uid, value, latitude, longitude, picture = row

        # Формируем словарь для сериализации
        data = {
            "name": name,
            "uid": uid,
            "value": value,
            "latitude": latitude,
            "longitude": longitude,
            "picture": picture,
        }

        # Отправляем данные на сериализацию
        serializer = CollectibleItemSerializer(data=data)

        # Проверяем валидность данных, если данные валидны, сохраняем их в базе данных
        if serializer.is_valid():
            # Сохраняем данные в базе данных
            serializer.save()
        else:
            # Если данные не валидны, добавляем их в список ошибок
            data_error.append(list(row))

    return data_error
